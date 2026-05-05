"""
Builds the output Excel file.

Columns: Fig | Q | C1 | C2 | C3 | C4 | A | Text Error | Error Answer
"""

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class ExcelRow:
    question_number: int
    question_text: str
    c1: str
    c2: str
    c3: str
    c4: str
    answer: str
    figure_path: str | None = None   # local path to figure image
    text_error: str = ""             # flagged OCR/extraction issues
    error_answer: str = ""           # flagged answer issues


# ── Styles ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="366092")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_CELL_FONT = Font(name="Calibri", size=10)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
_ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")

_HEADERS = ["Fig", "Q", "C1", "C2", "C3", "C4", "A", "Text Error", "Error Answer"]
_COL_WIDTHS = [18, 60, 30, 30, 30, 30, 20, 30, 30]
_ROW_HEIGHT = 80  # px — tall enough for an embedded figure thumbnail


def build_excel(rows: list[ExcelRow], output_path: str) -> str:
    """Write rows to Excel and return the saved file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Questions"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = _ROW_HEIGHT

        values = [
            "",                     # Fig — image inserted separately
            row.question_text,
            row.c1, row.c2, row.c3, row.c4,
            row.answer,
            row.text_error,
            row.error_answer,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _CELL_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER if col_idx in (1, 7) else _WRAP

            if col_idx in (8, 9) and value:
                cell.fill = _ERROR_FILL

        # Embed figure thumbnail in column A if available
        if row.figure_path and Path(row.figure_path).exists():
            try:
                img = XLImage(row.figure_path)
                img.width = 100
                img.height = 80
                cell_ref = f"A{row_idx}"
                ws.add_image(img, cell_ref)
            except Exception:
                ws.cell(row=row_idx, column=1, value="[image]")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
